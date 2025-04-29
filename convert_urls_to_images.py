import pandas as pd
import requests
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import argparse  # Import argparse module

import os
import subprocess
import time
from io import BytesIO
import re

def resize_image(img, max_width):
    """
    调整图片尺寸，同时保持图片质量
    """
    width, height = img.size
    print(f"Original Image Size: Width={width}, Height={height}")
    
    # 对于较小的图片，不进行缩放以保持清晰度
    if width <= max_width:
        return img, width, height
        
    # 计算高宽比
    ratio = max_width / width
    new_width = int(width * ratio)
    new_height = int(height * ratio)
    
    # 使用高质量的缩放方法
    # LANCZOS可能导致模糊，改用BICUBIC或保持原图
    img = img.resize((new_width, new_height), Image.BICUBIC)
    print(f"Resized Image Size: Width={new_width}, Height={new_height}")
    return img, new_width, new_height

def try_direct_conversion(url, output_path):
    """尝试直接从URL获取高质量图片版本"""
    # 尝试获取最高质量的版本
    try:
        # 如果URL包含 imageView2 参数，调整为获取原始或高质量版本
        if 'imageView2' in url:
            # 尝试获取原图（去除缩放参数）
            original_url = re.sub(r'imageView2/\d+/[^&]+', '', url)
            original_url = re.sub(r'&format/[^&]+', '', original_url)
            
            response = requests.get(original_url)
            response.raise_for_status()
            
            # 保存图片
            with open(output_path, 'wb') as f:
                f.write(response.content)
            
            # 验证是否为有效图片
            img = Image.open(output_path)
            img.verify()
            print(f"获取到高质量原图: {original_url}")
            return True
        
        # 如果URL包含 format/jpg 参数，尝试获取高质量版本
        if 'format/jpg' in url or 'format/jpeg' in url:
            # 尝试添加质量参数
            quality_url = url
            if 'q/' not in url:  # 如果没有质量参数
                quality_url = url + "&q/100"  # 添加最高质量参数
            
            response = requests.get(quality_url)
            response.raise_for_status()
            
            # 保存图片
            with open(output_path, 'wb') as f:
                f.write(response.content)
            
            # 验证是否为有效图片
            img = Image.open(output_path)
            img.verify()
            print(f"获取到高质量JPG: {quality_url}")
            return True
    except:
        pass
    
    # 尝试直接获取原始HEIC并转换（可能质量更高）
    try:
        response = requests.get(url)
        response.raise_for_status()
        
        # 保存HEIC文件
        heic_path = output_path + '.heic'
        with open(heic_path, 'wb') as f:
            f.write(response.content)
        
        # 使用高质量转换参数
        high_quality_conversion = [
            ['convert', '-quality', '100', heic_path, output_path],
            ['magick', '-quality', '100', heic_path, output_path],
            ['/usr/bin/convert', '-quality', '100', heic_path, output_path]
        ]
        
        for cmd in high_quality_conversion:
            try:
                subprocess.run(cmd, check=True)
                # 验证是否为有效图片
                img = Image.open(output_path)
                img.verify()
                os.remove(heic_path)  # 清理临时文件
                print(f"高质量转换成功: {cmd}")
                return True
            except:
                continue
        
        # 清理临时文件
        if os.path.exists(heic_path):
            os.remove(heic_path)
    except:
        pass
    
    return False

def download_and_insert_images(input_path, output_path, progress_callback=None):
    """
    下载并插入图片到Excel
    
    参数:
        input_path: 输入Excel文件路径
        output_path: 输出Excel文件路径
        progress_callback: 进度回调函数，接收当前进度和总进度参数
    """
    # Load the workbook
    wb = load_workbook(input_path)
    
    # Create a temporary directory to save images
    temp_dir = 'temp_images'
    os.makedirs(temp_dir, exist_ok=True)

    # 先计算需要处理的图片总数
    total_images = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        df = pd.DataFrame(ws.values)
        for row_idx, row in df.iterrows():
            for col_idx, cell in enumerate(row):
                if isinstance(cell, str) and '.heic' in cell:
                    total_images += 1
    
    # 处理进度计数
    processed_images = 0

    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        df = pd.DataFrame(ws.values)

        # Iterate through all cells to find HEIC URLs
        for row_idx, row in df.iterrows():
            for col_idx, cell in enumerate(row):
                if isinstance(cell, str) and '.heic' in cell:
                    url = cell
                    try:
                        # 更新进度和状态消息
                        progress_message = f"处理第 {processed_images+1}/{total_images} 张图片：{url[:30]}..."
                        if progress_callback:
                            progress_callback(processed_images, total_images, progress_message)
                        print(progress_message)
                        
                        # Clear the original link cell
                        ws.cell(row=row_idx + 4, column=col_idx + 1, value=None)
                        
                        # Output PNG path
                        png_path = os.path.join(temp_dir, f'image_{row_idx}_{col_idx}.png')
                        
                        # 尝试直接获取高质量图片版本
                        high_quality_success = try_direct_conversion(url, png_path)
                        
                        if not high_quality_success:
                            # 需要下载HEIC并转换
                            response = requests.get(url)
                            response.raise_for_status()

                            # Save the HEIC file to a temporary file
                            heic_path = os.path.join(temp_dir, f'image_{row_idx}_{col_idx}.heic')
                            with open(heic_path, 'wb') as f:
                                f.write(response.content)

                            # 更新进度消息
                            if progress_callback:
                                progress_callback(processed_images, total_images, f"转换图片 {processed_images+1}/{total_images}...")

                            # Convert HEIC to PNG using ImageMagick with high quality settings
                            conversion_success = False
                            
                            # 尝试不同的高质量转换命令
                            convert_commands = [
                                ['convert', '-quality', '100', heic_path, png_path],
                                ['magick', '-quality', '100', heic_path, png_path],
                                ['/usr/bin/convert', '-quality', '100', heic_path, png_path],
                                ['convert', heic_path, png_path],
                                ['magick', heic_path, png_path]
                            ]
                            
                            for cmd in convert_commands:
                                try:
                                    print(f"Trying conversion with: {' '.join(cmd)}")
                                    # 限制执行时间为60秒
                                    process = subprocess.Popen(cmd)
                                    start_time = time.time()
                                    while process.poll() is None and time.time() - start_time < 60:
                                        time.sleep(0.5)
                                    
                                    if process.poll() is None:
                                        # 进程还在运行，杀掉它
                                        process.terminate()
                                        process.wait(timeout=5)
                                        raise TimeoutError("Conversion timeout after 60 seconds")
                                    
                                    if process.returncode == 0:
                                        conversion_success = True
                                        break
                                except Exception as e:
                                    print(f"Command {cmd} failed with error: {e}")
                            
                            if not conversion_success:
                                raise Exception("All conversion methods failed")

                        # 更新进度消息
                        if progress_callback:
                            progress_callback(processed_images, total_images, f"调整图片尺寸 {processed_images+1}/{total_images}...")

                        # Open the converted PNG file using Pillow and resize if necessary
                        img = Image.open(png_path)
                        
                        # 调整图片尺寸，最大宽度为300像素，但保持质量
                        target_width = min(300, img.size[0])  # 如果原图已经小于300像素，保持原尺寸
                        img, img_width, img_height = resize_image(img, max_width=target_width)
                        
                        # 提高保存质量
                        img.save(png_path, quality=95, optimize=True)

                        # 更新进度消息
                        if progress_callback:
                            progress_callback(processed_images, total_images, f"插入图片到Excel {processed_images+1}/{total_images}...")

                        # Insert the image into Excel
                        img_for_excel = ExcelImage(png_path)
                        offset_col_idx = col_idx + 3  # Adjusted column index for offset
                        cell_ref = f"{chr(65 + offset_col_idx)}{row_idx + 1}"  # Convert adjusted index to Excel cell reference
                        ws.add_image(img_for_excel, cell_ref)
                        
                        # Adjust row height and column width if the image dimensions exceed current settings
                        img_height_points = img_height * 0.75  # 1 pixel is approximately 0.75 points
                        if ws.row_dimensions[row_idx + 1].height is None or ws.row_dimensions[row_idx + 1].height < img_height_points:
                            ws.row_dimensions[row_idx + 1].height = img_height_points

                        img_width_points = img_width * 0.14  # 1 pixel is approximately 0.14 column width units
                        col_letter = chr(65 + offset_col_idx)  # Adjusted column letter for offset
                        if ws.column_dimensions[col_letter].width is None or ws.column_dimensions[col_letter].width < img_width_points:
                            ws.column_dimensions[col_letter].width = img_width_points
                            
                        # 更新已处理图片数
                        processed_images += 1
                        
                        # 更新总体进度
                        if progress_callback:
                            progress_callback(processed_images, total_images, f"已完成 {processed_images}/{total_images} 张图片")

                    except requests.RequestException as e:
                        print(f"Failed to download image: {url}, Error: {e}")
                        processed_images += 1  # 虽然失败，但计为已处理
                    except subprocess.CalledProcessError as e:
                        print(f"ImageMagick conversion failed: {url}, Error: {e}")
                        processed_images += 1
                    except TimeoutError as e:
                        print(f"Conversion timeout: {url}, Error: {e}")
                        processed_images += 1
                    except Exception as e:
                        print(f"Failed to process image: {url}, Error: {e}")
                        processed_images += 1

    # 更新保存进度
    if progress_callback:
        progress_callback(processed_images, total_images, "正在保存Excel文件...")
    
    # Save the workbook
    wb.save(output_path)
    print(f"Saved Excel to: {output_path}")

    # 更新清理进度
    if progress_callback:
        progress_callback(total_images, total_images, "清理临时文件...")
        
    # Clean up the temporary directory
    for file in os.listdir(temp_dir):
        os.remove(os.path.join(temp_dir, file))
    os.rmdir(temp_dir)
    
    # 完成所有处理
    if progress_callback:
        progress_callback(total_images, total_images, "处理完成！")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Convert .heic links in Excel to images and insert them into the Excel file')
    parser.add_argument('input_path', type=str, help='Path to the input Excel file')
    parser.add_argument('output_path', type=str, help='Path to the output Excel file')

    args = parser.parse_args()

    download_and_insert_images(args.input_path, args.output_path)
